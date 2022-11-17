from openpyxl import Workbook
wb = Workbook()

ws = wb.active

ws['A1'] = "InvoiceNumber"
ws['B1'] = "stockcode"
ws['C1'] = "white hanging heart"
ws['D1'] = "Quantity"
ws['E1'] = "InvoiceDate"
ws['F1'] = "Unitprice"
ws['G1'] = "Customer ID"
ws['H1'] = "Country"

ws.append(['1234567','17221','white hanging heart','500gram','22-10-22','2.5','1028','INDIA'])


wb.save("Retail.xlsx")
